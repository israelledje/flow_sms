from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class Command(BaseCommand):
    help = 'Crée un fichier Excel d\'exemple pour l\'import des contacts'

    def handle(self, *args, **kwargs):
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"

        # Définir les en-têtes
        headers = [
            'phone_number',
            'first_name',
            'last_name',
            'email'
        ]

        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # Ajouter les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ajouter des exemples de données
        example_data = [
            ('237612345678', 'John', 'Doe', 'john.doe@example.com'),
            ('237612345679', 'Jane', 'Smith', 'jane.smith@example.com'),
            ('237612345680', 'Alice', 'Johnson', 'alice.j@example.com'),
            ('237612345681', 'Bob', 'Wilson', 'bob.wilson@example.com'),
            ('237612345682', 'Emma', 'Brown', 'emma.brown@example.com'),
        ]

        # Style pour les données
        data_alignment = Alignment(horizontal="left")

        # Ajouter les données d'exemple
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = data_alignment

        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Ajouter une note explicative
        ws.cell(row=len(example_data) + 2, column=1, value="Note :")
        ws.cell(row=len(example_data) + 2, column=2, value="Le numéro de téléphone doit commencer par '237' suivi de 8 ou 9 chiffres")
        ws.cell(row=len(example_data) + 3, column=1, value="Exemple :")
        ws.cell(row=len(example_data) + 3, column=2, value="237612345678")

        # Sauvegarder le fichier
        wb.save('contacts_exemple.xlsx')
        self.stdout.write(self.style.SUCCESS('Fichier Excel d\'exemple créé avec succès : contacts_exemple.xlsx')) 