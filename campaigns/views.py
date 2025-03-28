from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse
from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views import View
from .models import Campaign, Contact, SenderId, Message, SMSTemplate, ContactGroup
from django.db.models import Count, Q
import pandas as pd
import phonenumbers
from io import BytesIO
from django.utils import timezone
from django.contrib import messages
import re
import json
import csv
from openpyxl import Workbook
from django.core.exceptions import ValidationError
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class SenderIdListView(LoginRequiredMixin, ListView):
    model = SenderId
    template_name = 'campaigns/sender_id_list.html'
    context_object_name = 'sender_ids'

    def get_queryset(self):
        return SenderId.objects.filter(user=self.request.user)

class SenderIdCreateView(LoginRequiredMixin, CreateView):
    model = SenderId
    template_name = 'campaigns/sender_id_form.html'
    fields = ['name']
    success_url = reverse_lazy('campaigns:sender_id_list')

    def form_valid(self, form):
        form.instance.user = self.request.user
        # Vérifier si un sender ID avec le même nom existe déjà
        name = form.cleaned_data['name']
        existing_sender = SenderId.objects.filter(name=name).exists()
        
        # Vérifier aussi dans l'app user
        try:
            from django.apps import apps
            User_SenderId = apps.get_model('user', 'SenderID')
            existing_user_sender = User_SenderId.objects.filter(name=name).exists()
        except Exception:
            existing_user_sender = False
            
        if existing_sender or existing_user_sender:
            # Si le sender ID existe déjà, ajouter un message d'erreur au contexte
            form.add_error('name', 'Un expéditeur avec ce nom existe déjà')
            return self.form_invalid(form)
            
        return super().form_valid(form)

class ContactListView(LoginRequiredMixin, ListView):
    model = Contact
    template_name = 'campaigns/contact_list.html'
    context_object_name = 'contacts'

    def get_queryset(self):
        return Contact.objects.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contact_groups'] = ContactGroup.objects.filter(user=self.request.user)
        return context

class ContactCreateView(LoginRequiredMixin, CreateView):
    model = Contact
    template_name = 'campaigns/contact_form.html'
    fields = ['phone_number', 'first_name', 'last_name', 'email', 'groups']
    success_url = reverse_lazy('campaigns:contact_list')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            messages.success(self.request, "Contact créé avec succès!")
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, str(e))
            return self.form_invalid(form)

class ContactUpdateView(LoginRequiredMixin, UpdateView):
    model = Contact
    template_name = 'campaigns/contact_form.html'
    fields = ['phone_number', 'first_name', 'last_name', 'email', 'groups']
    success_url = reverse_lazy('campaigns:contact_list')

    def get_queryset(self):
        return Contact.objects.filter(user=self.request.user)

    def form_valid(self, form):
        try:
            self.object = form.save()
            messages.success(self.request, "Contact mis à jour avec succès!")
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, str(e))
            return self.form_invalid(form)

class CampaignListView(LoginRequiredMixin, ListView):
    model = Campaign
    template_name = 'campaigns/campaign_list.html'
    context_object_name = 'campaigns'

    def get_queryset(self):
        return Campaign.objects.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['draft_count'] = self.get_queryset().filter(status='draft').count()
        context['sending_count'] = self.get_queryset().filter(status='sending').count()
        context['sent_count'] = self.get_queryset().filter(status='sent').count()
        return context

class CampaignCreateView(LoginRequiredMixin, CreateView):
    model = Campaign
    template_name = 'campaigns/campaign_form.html'
    fields = ['name', 'message_content', 'sender', 'scheduled_date', 'is_rich_sms', 'rich_content']

    def get_success_url(self):
        return reverse('campaigns:campaign_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Récupérer les sender IDs depuis l'application user
        try:
            from django.apps import apps
            User_SenderId = apps.get_model('user', 'SenderID')
            
            # Si l'utilisateur est admin, il peut voir tous les Sender IDs approuvés
            if user.is_staff or user.is_superuser:
                user_approved_senders = User_SenderId.objects.filter(status="APPROUVE")
            else:
                # Sinon, l'utilisateur ne peut voir que ses propres Sender IDs approuvés
                user_approved_senders = User_SenderId.objects.filter(
                    user=user,
                    status="APPROUVE"
                )
            
            # Récupérer ou créer les SenderId de l'application campaigns correspondants
            campaign_senders = []
            for user_sender in user_approved_senders:
                campaign_sender, created = SenderId.objects.get_or_create(
                    user_sender_id=user_sender,
                    defaults={
                        'name': user_sender.name,
                        'user': user,
                        'status': 'approved'
                    }
                )
                campaign_senders.append(campaign_sender)
            
            context['senders'] = campaign_senders
                
        except Exception as e:
            print(f"DEBUG - Erreur en essayant de récupérer les Sender IDs depuis l'app user: {str(e)}")
            if user.is_staff or user.is_superuser:
                context['senders'] = SenderId.objects.filter(status='approved')
            else:
                context['senders'] = SenderId.objects.filter(
                    user=user,
                    status='approved'
                )
        
        # Ajouter les groupes de contacts au contexte
        context['contact_groups'] = ContactGroup.objects.filter(user=user)
        
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        
        # Sauvegarde la campagne
        self.object = form.save()
        
        # Récupérer et traiter les numéros de contact
        contact_numbers = self.request.POST.get('contact_numbers', '')
        if contact_numbers:
            # Créer ou récupérer les contacts et les associer à la campagne
            numbers = [num.strip() for num in contact_numbers.split(',') if num.strip()]
            for number in numbers:
                # Validation basique du numéro
                if not number or not re.match(r'^237[0-9]{8,9}$', number):
                    continue
                
                # Créer ou récupérer le contact
                contact, created = Contact.objects.get_or_create(
                    phone_number=number,
                    user=self.request.user
                )
                
                # Associer le contact à la campagne via le modèle Message
                Message.objects.create(
                    campaign=self.object,
                    contact=contact,
                    content=self.object.message_content,
                    is_rich_sms=self.object.is_rich_sms,
                    rich_content=self.object.rich_content,
                    sender=self.object.sender
                )
        
        # Gérer l'action selon le bouton cliqué
        action = self.request.POST.get('action', 'save')
        if action == 'start':
            try:
                campaign = self.object
                
                # Si la date planifiée est dans le futur, changer le statut à 'scheduled'
                if campaign.scheduled_date and campaign.scheduled_date > timezone.now():
                    campaign.status = 'scheduled'
                    campaign.save()
                    messages.success(self.request, f"La campagne '{campaign.name}' a été programmée pour {campaign.scheduled_date}.")
                else:
                    # Sinon, démarrer la campagne immédiatement
                    campaign.status = 'sending'
                    campaign.save()
                    campaign.send()
                    messages.success(self.request, f"La campagne '{campaign.name}' a été démarrée avec succès.")
            except Exception as e:
                messages.error(self.request, f"Erreur lors du démarrage de la campagne: {str(e)}")
        else:
            # Même en mode "save", vérifier si une date est programmée
            if self.object.scheduled_date and self.object.scheduled_date > timezone.now():
                self.object.status = 'scheduled'
                self.object.save()
                messages.success(self.request, f"La campagne '{self.object.name}' a été programmée pour {self.object.scheduled_date}.")
            else:
                messages.success(self.request, f"La campagne '{self.object.name}' a été enregistrée en tant que brouillon.")
        
        return redirect('campaigns:campaign_list')
        
    def form_invalid(self, form):
        # Ajouter un message pour aider l'utilisateur à comprendre pourquoi le formulaire n'est pas valide
        errors = form.errors.as_data()
        for field, error_list in errors.items():
            for error in error_list:
                messages.error(self.request, f"Erreur: {field} - {error.message}")
        
        return super().form_invalid(form)

class CampaignDetailView(LoginRequiredMixin, DetailView):
    model = Campaign
    template_name = 'campaigns/campaign_detail.html'
    context_object_name = 'campaign'

    def get_queryset(self):
        return Campaign.objects.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        messages = Message.objects.filter(campaign=self.object)
        context['messages'] = messages
        context['stats'] = {
            'total': messages.count(),
            'sent': messages.filter(status='sent').count(),
            'delivered': messages.filter(status='delivered').count(),
            'failed': messages.filter(status='failed').count(),
        }
        return context

class CampaignUpdateView(LoginRequiredMixin, UpdateView):
    model = Campaign
    template_name = 'campaigns/campaign_form.html'
    fields = ['name', 'message_content', 'sender', 'scheduled_date', 'is_rich_sms', 'rich_content']

    def get_success_url(self):
        return reverse('campaigns:campaign_list')

    def get_queryset(self):
        return Campaign.objects.filter(user=self.request.user, status='draft')

    def form_valid(self, form):
        # Sauvegarde la campagne
        self.object = form.save()
        
        # Supprimer les messages existants pour cette campagne
        Message.objects.filter(campaign=self.object).delete()
        
        # Récupérer et traiter les numéros de contact
        contact_numbers = self.request.POST.get('contact_numbers', '')
        if contact_numbers:
            # Créer ou récupérer les contacts et les associer à la campagne
            numbers = [num.strip() for num in contact_numbers.split(',') if num.strip()]
            for number in numbers:
                # Validation basique du numéro
                if not number or not re.match(r'^237[0-9]{8,9}$', number):
                    continue
                
                # Créer ou récupérer le contact
                contact, created = Contact.objects.get_or_create(
                    phone_number=number,
                )
                
                # Associer le contact à la campagne via le modèle Message
                Message.objects.create(
                    campaign=self.object,
                    contact=contact,
                    content=self.object.message_content,
                    is_rich_sms=self.object.is_rich_sms,
                    rich_content=self.object.rich_content,
                    sender=self.object.sender
                )
        
        # Gérer l'action selon le bouton cliqué
        action = self.request.POST.get('action', 'save')
        if action == 'start':
            try:
                campaign = self.object
                
                # Si la date planifiée est dans le futur, changer le statut à 'scheduled'
                if campaign.scheduled_date and campaign.scheduled_date > timezone.now():
                    campaign.status = 'scheduled'
                    campaign.save()
                    messages.success(self.request, f"La campagne '{campaign.name}' a été programmée pour {campaign.scheduled_date}.")
                else:
                    # Sinon, démarrer la campagne immédiatement
                    campaign.status = 'sending'
                    campaign.save()
                    campaign.send()
                    messages.success(self.request, f"La campagne '{campaign.name}' a été démarrée avec succès.")
            except Exception as e:
                messages.error(self.request, f"Erreur lors du démarrage de la campagne: {str(e)}")
        else:
            # Même en mode "save", vérifier si une date est programmée
            if self.object.scheduled_date and self.object.scheduled_date > timezone.now():
                self.object.status = 'scheduled'
                self.object.save()
                messages.success(self.request, f"La campagne '{self.object.name}' a été programmée pour {self.object.scheduled_date}.")
            else:
                messages.success(self.request, f"La campagne '{self.object.name}' a été mise à jour.")
        
        return redirect('campaigns:campaign_list')
        
    def form_invalid(self, form):
        # Ajouter un message pour aider l'utilisateur à comprendre pourquoi le formulaire n'est pas valide
        errors = form.errors.as_data()
        for field, error_list in errors.items():
            for error in error_list:
                messages.error(self.request, f"Erreur: {field} - {error.message}")
        
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Récupérer les sender IDs depuis l'application user
        try:
            from django.apps import apps
            User_SenderId = apps.get_model('user', 'SenderID')
            
            # Si l'utilisateur est admin, il peut voir tous les Sender IDs approuvés
            if user.is_staff or user.is_superuser:
                user_approved_senders = User_SenderId.objects.filter(status="APPROUVE")
                print(f"DEBUG - Admin - Tous les Sender IDs approuvés: {list(user_approved_senders.values('id', 'name', 'user__username'))}")
            else:
                # Sinon, l'utilisateur ne peut voir que ses propres Sender IDs approuvés
                user_approved_senders = User_SenderId.objects.filter(
                    user=user,
                    status="APPROUVE"
                )
                print(f"DEBUG - Utilisateur - Sender IDs approuvés: {list(user_approved_senders.values('id', 'name'))}")
            
            # Récupérer ou créer les SenderId de l'application campaigns correspondants
            campaign_senders = []
            for user_sender in user_approved_senders:
                # Vérifier si un SenderId correspondant existe déjà
                campaign_sender, created = SenderId.objects.get_or_create(
                    user_sender_id=user_sender,
                    defaults={
                        'name': user_sender.name,
                        'user': user,
                        'status': 'approved'  # Status dans campaigns.SenderId
                    }
                )
                campaign_senders.append(campaign_sender)
            
            context['senders'] = campaign_senders
                
        except Exception as e:
            print(f"DEBUG - Erreur en essayant de récupérer les Sender IDs depuis l'app user: {str(e)}")
            # Fallback sur les Sender IDs de l'application campaigns
            if user.is_staff or user.is_superuser:
                context['senders'] = SenderId.objects.filter(status='approved')
            else:
                context['senders'] = SenderId.objects.filter(
                    user=user,
                    status='approved'
                )
        
        # Vérifier les variables de user pour débogage
        print(f"DEBUG - User ID: {user.id}, Username: {user.username}, Is staff: {user.is_staff}, Is superuser: {user.is_superuser}")
        
        return context

class CampaignDeleteView(LoginRequiredMixin, DeleteView):
    model = Campaign
    template_name = 'campaigns/campaign_confirm_delete.html'
    success_url = reverse_lazy('campaigns:campaign_list')

    def get_queryset(self):
        return Campaign.objects.filter(user=self.request.user, status='draft')

class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'campaigns/reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        campaigns = Campaign.objects.filter(user=self.request.user)
        
        context['stats'] = {
            'total_campaigns': campaigns.count(),
            'total_messages': Message.objects.filter(campaign__in=campaigns).count(),
            'delivery_rate': Message.objects.filter(
                campaign__in=campaigns,
                status='delivered'
            ).count() / Message.objects.filter(campaign__in=campaigns).count() * 100 if Message.objects.filter(campaign__in=campaigns).exists() else 0,
            'campaigns_by_status': campaigns.values('status').annotate(count=Count('id'))
        }
        
        return context

class CampaignSendView(LoginRequiredMixin, View):
    def post(self, request, pk):
        campaign = get_object_or_404(Campaign, pk=pk, user=request.user)
        try:
            campaign.send()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

class ContactGroupListView(LoginRequiredMixin, ListView):
    model = ContactGroup
    template_name = 'campaigns/contact_group_list.html'
    context_object_name = 'contact_groups'

    def get_queryset(self):
        return ContactGroup.objects.filter(user=self.request.user)

class ContactGroupCreateView(LoginRequiredMixin, CreateView):
    model = ContactGroup
    template_name = 'campaigns/contact_group_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('campaigns:contact_group_list')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

class ContactGroupUpdateView(LoginRequiredMixin, UpdateView):
    model = ContactGroup
    template_name = 'campaigns/contact_group_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('campaigns:contact_group_list')

    def get_queryset(self):
        return ContactGroup.objects.filter(user=self.request.user)

class ContactGroupDeleteView(LoginRequiredMixin, DeleteView):
    model = ContactGroup
    template_name = 'campaigns/contact_group_confirm_delete.html'
    success_url = reverse_lazy('campaigns:contact_group_list')

    def get_queryset(self):
        return ContactGroup.objects.filter(user=self.request.user)

class ContactExportView(LoginRequiredMixin, View):
    def get(self, request):
        format_type = request.GET.get('format', 'csv')
        contacts = Contact.objects.filter(user=request.user)
        
        if format_type == 'excel':
            # Créer un nouveau classeur Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Contacts"
            
            # En-têtes
            headers = ['Prénom', 'Nom', 'Numéro de téléphone', 'Email', 'Groupes']
            ws.append(headers)
            
            # Données
            for contact in contacts:
                ws.append([
                    contact.first_name or '',
                    contact.last_name or '',
                    contact.phone_number,
                    contact.email or '',
                    contact.group_names
                ])
            
            # Créer la réponse
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="contacts.xlsx"'
            wb.save(response)
            return response
            
        else:  # CSV par défaut
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="contacts.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Prénom', 'Nom', 'Numéro de téléphone', 'Email', 'Groupes'])
            
            for contact in contacts:
                writer.writerow([
                    contact.first_name or '',
                    contact.last_name or '',
                    contact.phone_number,
                    contact.email or '',
                    contact.group_names
                ])
            
            return response

class GetGroupContactsView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            group_ids = data.get('groups', [])
            
            contacts = Contact.objects.filter(
                user=request.user,
                groups__id__in=group_ids,
                is_active=True
            ).distinct().values('id', 'phone_number')
            
            return JsonResponse({
                'status': 'success',
                'contacts': list(contacts)
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)

class ContactImportView(LoginRequiredMixin, View):
    def post(self, request):
        if 'file' not in request.FILES:
            return JsonResponse({
                'status': 'error', 
                'message': 'Aucun fichier fourni'
            }, status=400)
        
        contact_file = request.FILES['file']
        group_id = request.POST.get('group')
        
        try:
            # Lecture du fichier
            if contact_file.name.endswith('.csv'):
                df = pd.read_csv(contact_file, dtype={'phone_number': str})
            else:
                df = pd.read_excel(contact_file, dtype={'phone_number': str})
            
            # Validation des colonnes
            if 'phone_number' not in df.columns:
                return JsonResponse({
                    'status': 'error',
                    'message': 'La colonne "phone_number" est requise'
                }, status=400)
            
            # Traitement des contacts
            valid_contacts = []
            invalid_numbers = []
            duplicates = []
            
            for _, row in df.iterrows():
                phone = str(row['phone_number']).strip()
                
                # Nettoyage du numéro (suppression des espaces et caractères spéciaux)
                phone = re.sub(r'[^0-9]', '', phone)
                
                # Si le numéro ne commence pas par 237, l'ajouter
                if not phone.startswith('237') and len(phone) in [8, 9]:
                    phone = '237' + phone
                
                # Vérification du format du numéro (237 + 8 ou 9 chiffres)
                if not re.match(r'^237[0-9]{8,9}$', phone):
                    print(f"Numéro invalide: {phone}")
                    invalid_numbers.append(phone)
                    continue
                
                print(f"Numéro valide: {phone}")
                
                # Vérification des doublons
                if Contact.objects.filter(user=request.user, phone_number=phone).exists():
                    duplicates.append(phone)
                    continue
                
                # Création du contact
                contact_data = {
                    'phone_number': phone,
                    'user': request.user,
                }
                
                # Ajout des champs optionnels
                for field in ['first_name', 'last_name', 'email']:
                    if field in df.columns and pd.notna(row[field]):
                        contact_data[field] = str(row[field]).strip()
                
                # Création du contact sans validation supplémentaire
                valid_contacts.append(Contact(**contact_data))
            
            # Sauvegarde des contacts valides
            Contact.objects.bulk_create(valid_contacts)
            
            # Ajout au groupe si spécifié
            if group_id and valid_contacts:
                group = ContactGroup.objects.get(id=group_id, user=request.user)
                contacts_to_add = Contact.objects.filter(
                    user=request.user,
                    phone_number__in=[c.phone_number for c in valid_contacts]
                )
                group.contacts.add(*contacts_to_add)
            
            return JsonResponse({
                'status': 'success',
                'data': {
                    'valid_count': len(valid_contacts),
                    'invalid_count': len(invalid_numbers),
                    'duplicate_count': len(duplicates),
                    'total_count': len(df)
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Erreur lors du traitement du fichier: {str(e)}'
            }, status=500)

class SMSTemplateListView(LoginRequiredMixin, ListView):
    model = SMSTemplate
    template_name = 'campaigns/sms_template_list.html'
    context_object_name = 'templates'

    def get_queryset(self):
        return SMSTemplate.objects.filter(user=self.request.user)

class SMSTemplateCreateView(LoginRequiredMixin, CreateView):
    model = SMSTemplate
    template_name = 'campaigns/sms_template_form.html'
    fields = ['name', 'content']
    success_url = reverse_lazy('campaigns:sms_template_list')

    def form_valid(self, form):
        form.instance.user = self.request.user
        # Récupérer les variables du formulaire
        variables = self.request.POST.getlist('variables')
        # Filtrer les variables vides et les doublons
        variables = list(set(v.strip() for v in variables if v.strip()))
        form.instance.variables = variables
        return super().form_valid(form)

    def form_invalid(self, form):
        # Ajouter un message pour aider l'utilisateur à comprendre pourquoi le formulaire n'est pas valide
        errors = form.errors.as_data()
        for field, error_list in errors.items():
            for error in error_list:
                messages.error(self.request, f"Erreur: {field} - {error.message}")
        return super().form_invalid(form)

class SMSTemplateUpdateView(LoginRequiredMixin, UpdateView):
    model = SMSTemplate
    template_name = 'campaigns/sms_template_form.html'
    fields = ['name', 'content', 'variables']
    success_url = reverse_lazy('campaigns:sms_template_list')

    def get_queryset(self):
        return SMSTemplate.objects.filter(user=self.request.user)

class SMSTemplateDeleteView(LoginRequiredMixin, DeleteView):
    model = SMSTemplate
    template_name = 'campaigns/sms_template_confirm_delete.html'
    success_url = reverse_lazy('campaigns:sms_template_list')

    def get_queryset(self):
        return SMSTemplate.objects.filter(user=self.request.user)

class GetSMSTemplatesView(LoginRequiredMixin, View):
    def get(self, request):
        templates = SMSTemplate.objects.filter(user=request.user).values('id', 'name')
        return JsonResponse({'templates': list(templates)}, safe=False)

class GetTemplateDetailsView(LoginRequiredMixin, View):
    def get(self, request, template_id):
        template = get_object_or_404(SMSTemplate, id=template_id, user=request.user)
        return JsonResponse({
            'id': template.id,
            'name': template.name,
            'content': template.content,
            'variables': template.variables
        })

class GetTemplateVariablesView(LoginRequiredMixin, View):
    def get(self, request):
        template_id = request.GET.get('template_id')
        if template_id:
            template = get_object_or_404(SMSTemplate, id=template_id, user=request.user)
            return JsonResponse({'variables': template.variables})
        return JsonResponse({'variables': []})

class ContactDeleteView(LoginRequiredMixin, DeleteView):
    model = Contact
    template_name = 'campaigns/contact_confirm_delete.html'
    success_url = reverse_lazy('campaigns:contact_list')

    def get_queryset(self):
        return Contact.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Contact supprimé avec succès!")
        return super().delete(request, *args, **kwargs)

class DownloadTemplateView(LoginRequiredMixin, View):
    def get(self, request):
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"

        # Définir les en-têtes
        headers = [
            'phone_number',
            'first_name',
            'last_name',
            'email'
        ]

        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # Ajouter les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ajouter des exemples de données
        example_data = [
            ('237612345678', 'John', 'Doe', 'john.doe@example.com'),
            ('237612345679', 'Jane', 'Smith', 'jane.smith@example.com'),
            ('237612345680', 'Alice', 'Johnson', 'alice.j@example.com'),
            ('237612345681', 'Bob', 'Wilson', 'bob.wilson@example.com'),
            ('237612345682', 'Emma', 'Brown', 'emma.brown@example.com'),
        ]

        # Style pour les données
        data_alignment = Alignment(horizontal="left")

        # Ajouter les données d'exemple
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = data_alignment

        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Ajouter une note explicative
        ws.cell(row=len(example_data) + 2, column=1, value="Note :")
        ws.cell(row=len(example_data) + 2, column=2, value="Le numéro de téléphone doit commencer par '237' suivi de 8 ou 9 chiffres")
        ws.cell(row=len(example_data) + 3, column=1, value="Exemple :")
        ws.cell(row=len(example_data) + 3, column=2, value="237612345678")

        # Créer la réponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="contacts_template.xlsx"'
        wb.save(response)
        return response
